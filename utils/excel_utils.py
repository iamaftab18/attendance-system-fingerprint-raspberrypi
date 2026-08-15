import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from flask import current_app


def _get_excel_path(lecture):
    folder = current_app.config['ATTENDANCE_FOLDER']
    date_str = lecture.lecture_date.strftime('%Y%m%d')
    filename = f"Attendance_{lecture.subject.replace(' ', '_')}_{date_str}_L{lecture.id}.xlsx"
    return os.path.join(folder, filename), filename


def _header_style():
    fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    return fill, font, align, border


def _apply_row_style(ws, row_num, even=True):
    bg = 'EFF6FF' if even else 'DBEAFE'
    fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    font = Font(name='Calibri', size=10)
    align = Alignment(horizontal='center', vertical='center')
    for col in range(1, 8):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def create_lecture_excel(lecture):
    path, filename = _get_excel_path(lecture)

    wb = openpyxl.Workbook()

    # ── IN Sheet ──────────────────────────────────────────────────────────────
    ws_in = wb.active
    ws_in.title = 'IN Attendance'

    # Title block
    ws_in.merge_cells('A1:G1')
    title_cell = ws_in['A1']
    title_cell.value = 'ATTENDANCE MANAGEMENT SYSTEM'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='1E40AF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_in.row_dimensions[1].height = 30

    ws_in.merge_cells('A2:G2')
    sub_cell = ws_in['A2']
    sub_cell.value = f"Subject: {lecture.subject}  |  Date: {lecture.lecture_date.strftime('%d/%m/%Y')}  |  Time: {lecture.start_time} - {lecture.end_time}"
    sub_cell.font = Font(name='Calibri', size=10, color='374151')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    sub_cell.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    ws_in.row_dimensions[2].height = 22

    ws_in.merge_cells('A3:G3')
    dept_cell = ws_in['A3']
    dept_cell.value = f"Teacher: {lecture.teacher.name}  |  Department ID: {lecture.department_id}  |  Year: {lecture.year}"
    dept_cell.font = Font(name='Calibri', size=10, italic=True, color='6B7280')
    dept_cell.alignment = Alignment(horizontal='center')
    dept_cell.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    ws_in.row_dimensions[3].height = 20

    ws_in.row_dimensions[4].height = 6  # spacer

    # Headers
    headers = ['#', 'PRN', 'Student Name', 'In Time', 'Status', 'Face Image', 'Notes']
    fill, font, align, border = _header_style()
    for col, header in enumerate(headers, 1):
        cell = ws_in.cell(row=5, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws_in.row_dimensions[5].height = 24

    # Column widths
    col_widths = [5, 15, 28, 14, 12, 35, 20]
    for i, width in enumerate(col_widths, 1):
        ws_in.column_dimensions[get_column_letter(i)].width = width

    # ── OUT Sheet ─────────────────────────────────────────────────────────────
    ws_out = wb.create_sheet('OUT Attendance')
    ws_out.merge_cells('A1:G1')
    t2 = ws_out['A1']
    t2.value = 'OUT ATTENDANCE - ' + lecture.subject
    t2.font = Font(name='Calibri', bold=True, size=13, color='065F46')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    t2.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    ws_out.row_dimensions[1].height = 28

    ws_out.merge_cells('A2:G2')
    s2 = ws_out['A2']
    s2.value = f"Date: {lecture.lecture_date.strftime('%d/%m/%Y')}  |  Lecture End: {lecture.end_time}"
    s2.font = Font(name='Calibri', size=10, color='374151')
    s2.alignment = Alignment(horizontal='center')
    s2.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    ws_out.row_dimensions[2].height = 20

    ws_out.row_dimensions[3].height = 6

    out_headers = ['#', 'PRN', 'Student Name', 'In Time', 'Out Time', 'Duration (min)', 'Face Image']
    out_fill = PatternFill(start_color='065F46', end_color='065F46', fill_type='solid')
    for col, header in enumerate(out_headers, 1):
        cell = ws_out.cell(row=4, column=col, value=header)
        cell.fill = out_fill
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_out.row_dimensions[4].height = 24

    for i, width in enumerate(col_widths, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = width

    if 'OUT Attendance' in wb.sheetnames:
        del wb['OUT Attendance']

    wb.save(path)
    return filename


def append_student_to_excel(lecture, record):
    path, _ = _get_excel_path(lecture)
    if not os.path.exists(path):
        create_lecture_excel(lecture)

    wb = openpyxl.load_workbook(path)
    ws = wb['IN Attendance']

    # Find next row (after header rows 1-5)
    row = ws.max_row + 1
    if row < 6:
        row = 6

    idx = row - 5
    even = idx % 2 == 0

    student = record.student
    in_time = record.in_time.strftime('%H:%M:%S') if record.in_time else ''
    face_img = record.face_image_path or ''

    data = [idx, student.prn, student.name, in_time, 'Present', face_img, '']
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
    _apply_row_style(ws, row, even)
    ws.row_dimensions[row].height = 20

    wb.save(path)


def update_excel_out_time(lecture_id, record):
    from models import Lecture
    from flask import current_app
    lecture = Lecture.query.get(lecture_id)
    if not lecture:
        return

    path, _ = _get_excel_path(lecture)
    if not os.path.exists(path):
        return

    wb = openpyxl.load_workbook(path)

    # Update IN sheet status
    ws_in = wb['IN Attendance']
    for row in ws_in.iter_rows(min_row=6):
        if row[1].value == record.student.prn:
            row[4].value = 'Out'
            row[4].fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            break

    # Add to OUT sheet
    ws_out = wb['OUT Attendance']
    next_row = ws_out.max_row + 1
    if next_row < 5:
        next_row = 5

    student = record.student
    in_t = record.in_time.strftime('%H:%M:%S') if record.in_time else ''
    out_t = record.out_time.strftime('%H:%M:%S') if record.out_time else ''

    duration = ''
    if record.in_time and record.out_time:
        diff = record.out_time - record.in_time
        duration = str(round(diff.total_seconds() / 60, 1))

    out_data = [next_row - 4, student.prn, student.name, in_t, out_t, duration, record.face_image_path or '']
    for col, val in enumerate(out_data, 1):
        ws_out.cell(row=next_row, column=col, value=val)

    even = (next_row - 4) % 2 == 0
    out_fill_color = 'D1FAE5' if even else 'A7F3D0'
    out_fill = PatternFill(start_color=out_fill_color, end_color=out_fill_color, fill_type='solid')
    for col in range(1, 8):
        ws_out.cell(row=next_row, column=col).fill = out_fill
        ws_out.cell(row=next_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
    ws_out.row_dimensions[next_row].height = 20

    wb.save(path)


def finalize_lecture_excel(lecture):
    path, filename = _get_excel_path(lecture)
    if not os.path.exists(path):
        create_lecture_excel(lecture)

    wb = openpyxl.load_workbook(path)

    # Add summary sheet
    ws_sum = wb.create_sheet('Summary', 0)
    ws_sum.merge_cells('A1:D1')
    ws_sum['A1'].value = 'ATTENDANCE SUMMARY'
    ws_sum['A1'].font = Font(name='Calibri', bold=True, size=14, color='1E40AF')
    ws_sum['A1'].alignment = Alignment(horizontal='center')
    ws_sum['A1'].fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

    from models import AttendanceRecord
    records = AttendanceRecord.query.filter_by(lecture_id=lecture.id).all()
    present = sum(1 for r in records if r.status == 'present')
    total_students = len(records)

    summary_data = [
        ('Subject', lecture.subject),
        ('Date', lecture.lecture_date.strftime('%d/%m/%Y')),
        ('Start Time', lecture.start_time),
        ('End Time', lecture.end_time),
        ('Teacher', lecture.teacher.name),
        ('Total Present', present),
        ('Total Students Scanned', total_students),
        ('Generated At', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
    ]

    for i, (key, val) in enumerate(summary_data, 3):
        ws_sum.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=val)

    ws_sum.column_dimensions['A'].width = 25
    ws_sum.column_dimensions['B'].width = 35

    wb.save(path)
    return filename
